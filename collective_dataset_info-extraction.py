from openpyxl import Workbook, load_workbook

import os

import pandas as pd
import csv
import openpyxl


# groups all files in Datasets directory into an unordered list
data_files = os.listdir('.\\Datasets')
#print(data_files) 


# orders all files in order of numbered prefix, 1-, 2-, 3-, etc.
'''
ordered_files = []
next_file = 1
while len(ordered_files) < len(data_files):
    for file in data_files:
        #print(len(files), len(data_files))
        order_number = int(file.partition('-')[0])
        if order_number == next_file:
            ordered_files.append(file)
            next_file += 1
        else:
            pass
'''


# ordered_files is a list of file names in order:
'''
for i in ordered_files:
    print(i)
'''


# Collect CSV Headers
def csv_headers(file_path):
    with open(file_path) as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_headers = row
            break
    return csv_headers

# print(csv_headers('.\\Datasets\\1-Divvy_Trips_2013.csv'))


def csv_fields_records_count(file_path):
    with open(file_path) as file_obj:
        reader = csv.reader(file_obj)
        row_count = 0
        for row in reader:
            row_count += 1
        row_count -= 1
        column_count = len(csv_headers(file_path))
    return row_count, column_count

print(csv_fields_records_count('.\\Datasets\\0-TEST_TRIPS.csv'))





# NEW FUNCTION OR SECOND ATTEMPT:   
def csv_empty_field_record_check(file_path):
    '''
    If there are any empty fields or records within the
    csv file, then the function will return a Boolean
    tuple (empty_column, empty_row) and the number of
    fields and records that are empty. A True Boolean
    implies that an entire field or record is empty.
    '''
    number_of_headers = len(csv_headers(file_path))
    column_length = csv_fields_records_count(file_path)[0]
    empty_column_count = 0
    with open(file_path) as file_obj:
        reader = csv.reader(file_obj)
        next(reader) # skips 1st row or column headers
        for col_num in range(number_of_headers):
            column_entry_count = 0
            #print(empty_column_count) ###
            for row in reader:
                if row[col_num] != '':
                    print(col_num, row[col_num]) ### 
                    break #problem has something do with break
                elif row[col_num] == '':
                    column_entry_count += 1
                    print(col_num,'    ', row[col_num], '    ', column_entry_count) ###
                    if column_entry_count == column_length:
                        empty_column_count += 1
                    else:
                        pass
    return # empty_column_count
print(csv_empty_field_record_check('.\\Datasets\\0-TEST_TRIPS.csv'))

print('\n\n\n')

'''
with open('.\\Datasets\\0-TEST_TRIPS.csv') as file_obj:
    reader = csv.reader(file_obj)
    next(reader) # skips 1st row or column headers
    for row in reader:
        print(row)
'''







def csv_empty_field_record_check(file_path):
    '''
    If there are any empty fields or records within the
    csv file, then the function will return a Boolean
    tuple (empty_column, empty_row). A True Boolean
    implies that an entire field or record is empty.
    '''
    empty_column = 0
    number_of_headers = len(csv_headers(file_path))
    with open(file_path) as file_obj:
        reader = csv.reader(file_obj)
        next(reader) # skips row of headers
        for col_num in range(number_of_headers):
            if col_num == 7: ###
                print('COLUMN 8') ###
            # CHECKS FOR EMPTY COLUMNS
            for row in reader:
                if row[col_num] != '':
                    empty_column = False
                    break    
                elif row[col_num] == '': # if row entry = ''
                    empty_column = True
            if col_num == 7: ###
                print(empty_column) ###
                print(row[col_num]) ###
                print(row) ###

    print('\n\n\n\n')
    
    empty_row = 0
    empty_field_counter = 0
    with open(file_path) as file_obj: 
        reader = csv.reader(file_obj)
        next(reader) # skips row of headers
        for row in reader:
            # CHECKS FOR EMPTY ROWS 
            if empty_field_counter == number_of_headers:
                break
            for entry in row:
                if entry != '':
                    empty_row = False
                    break
                else: # if row entry = ''
                    empty_field_counter += 1
                    empty_row = True
            if col_num == 7: ###
                print(empty_column) ###
                print(row[col_num]) ###
                print(row) ###
                
        return empty_column, empty_row 
                
#print(csv_empty_field_record_check('.\\Datasets\\0-TEST_TRIPS.csv'))
# CHANGES ARE NECESSARY! Number of EMPTY columns and rows to be included also!

# make separate function for any missing fields in entire dataset














































def csv_empty_field_record_check(file_path):
    '''
    If there are any empty fields or records within the
    csv file, then the function will return a Boolean
    tuple (empty_column, empty_row). A True Boolean
    implies that an entire field or record is empty.
    '''
    empty_column = 0
    number_of_headers = len(csv_headers(file_path))
    with open(file_path) as file_obj:
        reader = csv.reader(file_obj)
        next(reader) # skips row of headers
        for col_num in range(number_of_headers):
            # CHECKS FOR EMPTY COLUMNS
            for row in reader:
                if row[col_num] != '':
                    empty_column = False
                    break    
                else: # if row entry = ''
                    empty_column = True
    empty_row = 0
    empty_field_counter = 0
    with open(file_path) as file_obj: 
        reader = csv.reader(file_obj)
        next(reader) # skips row of headers
        for row in reader:
            # CHECKS FOR EMPTY ROWS 
            if empty_field_counter == number_of_headers:
                break
            for entry in row:
                if entry != '':
                    empty_row = False
                    break
                else: # if row entry = ''
                    empty_field_counter += 1
                    empty_row = True
        return empty_column, empty_row 
                
# print(csv_empty_field_record_check('.\\Datasets\\0-TEST_TRIPS.csv'))
# CHANGES MAY BE NECESSARY! Number of EMPTY columns and rows to be included also!

# make separate function for any missing fields in entire dataset




# Extract complete records from CSV:

def csv_complete_records(file_path, number_of_records):
    '''
    Collects a specified number of complete-records, i.e., rows with
    no missing fields, within a csv file. Exception: If there exists
    an empty column in the csv file, then a 'complete-record' will
    still be returned with a missing field within it.
    '''
    number_of_headers = len(csv_headers(file_path))
    complete_records = []
    with open(file_path) as file_obj:
        reader = csv.reader(file_obj)
        next(reader)
        if csv_empty_field_record_check(file_path)[0] == False:
        # column condition ADDED: returns FALSE for empty column
            for row in reader:
                if len(complete_records) < number_of_records:
                    if '' in row: 
                        pass
                    else: # condition 1: NO empty strings in row
                        complete_records.append(row)
                    #print('LENGTH:', len(complete_records))
                else:
                    break
        else:
            pass #number_of_headers - 
    return complete_records

#print(csv_complete_records('.\\Datasets\\10-Divvy_Trips_2015-Q2.csv', 3))
# 1-Divvy_Trips_2013


















# Extract complete records from CSV-OLD:

def csv_complete_records(file_path, number_of_records):
    '''
    Collects a specified number of complete-records, i.e., rows with
    no missing fields, within a csv file. Exception: If there exists
    an empty column in the csv file, then a 'complete-record' will
    still be returned with a missing field within it.
    '''
    number_of_headers = len(csv_headers(file_path))
    complete_records = []
    with open(file_path) as file_obj:
        reader = csv.reader(file_obj)
        next(reader)
        # column condition that returns FALSE for empty column
        for row in reader:
            if len(complete_records) < number_of_records:
                if '' in row: 
                    pass
                else: # condition 1: NO empty strings in row
                    complete_records.append(row)
                #print('LENGTH:', len(complete_records))
            else:
                break
    return complete_records

#print(csv_complete_records('.\\Datasets\\10-Divvy_Trips_2015-Q2.csv', 10))
# 1-Divvy_Trips_2013











# use of pandas to save csv as excel - FASTEST SO FAR! MAKE IT FASTER!!!
# confirm number of csv records and fields do not exceed Excel's limit, i.e., 1,048,576 rows and 16,384 columns
# use CSV module to collect overview-data
# IF and ONLY if ALL CSV files do not exceed Excel's limit, then convert to xlsx extension files.
'''
df = pd.read_csv('.\\Datasets\\1-Divvy_Trips_2013.csv', dtype =
                 {'trip_id': 'str',
                  'starttime': 'str',
                  'stoptime': 'str',
                  'bikeid': 'str',
                  'tripduration': 'str',
                  'from_station_id': 'str',
                  'from_station_name': 'str',
                  'to_station_id': 'str',
                  'to_station_name': 'str',
                  'usertype': 'str',
                  'gender': 'str',
                  'birthday': 'str'})

writer = pd.ExcelWriter('TESTING_CSV_TO_EXCEL.xlsx')
df.to_excel(writer, index = False, header = True)
writer.save()
'''















# use of pandas to save csv as excel - SLOW!
'''
df = pd.read_csv('.\\Datasets\\1-Divvy_Trips_2013.csv', low_memory = False)
writer = pd.ExcelWriter('TESTING_CSV_TO_EXCEL.xlsx')
df.to_excel(writer, index = False, header = True)
writer.save()
'''


# use of pandas to save csv as excel - SLOW or NOT WORKING!
'''
csv = pd.read_csv('.\\Datasets\\1-Divvy_Trips_2013.csv')
read_file.to_excel('.\\Datasets\\1-Divvy_Trips_2013---EXCEL_COPY.xlsx', index = None, header = True)
'''


# use of csv module to copy into excel and save - SLOW!
'''
temp_book = Workbook()
sheet = temp_book.active
with open('.\\Datasets\\1-Divvy_Trips_2013.csv') as file_obj:
    reader = csv.reader(file_obj)
    for row in reader:
        sheet.append(row)
temp_book.save('Sample.xlsx')
'''











# temporarily convert csv file to xlsx format
'''
def csv_to_xlsx(csv_file, excel_spreadsheet, save_file = False):
    csv = pd.read_csv(csv_file)
    excelWriter = pd.ExcelWriter(excel_spreadsheet)
    csv.to_excel(excelWriter)
    if save_file == True:
        excelWriter.save()
    else:
        pass
'''




# Access and extract specifics:
'''
for data_file in ordered_files:
    #csv_to_xlsx(f'.\\Datasets\\{data_file}', 'temp_data_placeholder.xlsx')
    #wb = load_workbook(f'.\\Datasets\\{data_file}')
    print(f'.\\Datasets\\{data_file}')
'''








# Newly Created Summary xlsx File:
'''
headers = ['Dataset File', 'Timeframe of Data Collection', 'Record Count',
           'Field Count', 'Provision of Metadata File', 'Provision of Shape File',
           'Incomplete Dataset']

wb = Workbook()

wb['Sheet'].title = 'Dataset Specifics'

wb.create_sheet('Dataset Previews')

print(wb.sheetnames)


ws1 = wb['Dataset Specifics']
ws2 = wb['Dataset Previews']

print(ws1)
print(ws2)

ws1.append(headers)

wb.save('Collective_Data_Summary.xlsx')
'''









